import pandas as pd
from openpyxl.styles import Alignment
from app.internal.config.config_style import HEADER_FILL, HEADER_FONT, HEADER_FONT_RED, HEADERS
from app.internal.config.char import ILLEGAL_CHARS_RE


class FEC:
    def __init__(self, file_path: str, original_filename: str = ""):
        self.path = file_path
        self.original_filename = original_filename if original_filename else file_path
        self.df = pd.DataFrame()
        self.date = ""
        self.group_name = ""

    def open_fec(self) -> pd.DataFrame:
        encodings = ['utf-8-sig', 'latin-1', 'iso-8859-1', 'cp1252']

        for encoding in encodings:
            try:
                print(f"Tentative de lecture avec l'encodage: {encoding}")
                self.df = pd.read_csv(self.path, sep='\t', encoding=encoding)
                print(f"Succès avec l'encodage: {encoding}")
                break
            except UnicodeDecodeError:
                print(f"Échec avec l'encodage: {encoding}")
                continue
        else:
            raise ValueError("Impossible de lire le fichier avec les encodages testés")

        self.df.columns = self.df.columns.str.strip()
        self.df['JournalCode'] = self.df['JournalCode'].str.strip()

        print("Colonnes disponibles:", self.df.columns.tolist())
        print(self.df.head())

        return self.df

    def set_final_fec_date(self):
        filename = self.original_filename.split('/')[-1]  # ex: CATEST410217533FEC20251231.txt
        name_without_ext = filename.split('.')[0]  # ex: CATEST410217533FEC20251231
        raw = name_without_ext[-8:]  # ex: 20251231
        self.date = f"{raw[6:8]}/{raw[4:6]}/{raw[0:4]}"  # ex: 31/12/2025
        before_fec = name_without_ext.split('FEC')[0]  # ex: CATEST410217533
        self.group_name = before_fec[:-9]  # ex: CATEST (retire les 9 chiffres du SIREN)
        print(self.date, self.group_name)

    def create_entete(self, ws):
        ws["A1"].value = self.group_name
        ws["A1"].font = HEADER_FONT
        ws["A1"].fill = HEADER_FILL

        ws["A2"].value = self.date
        ws["A2"].font = HEADER_FONT
        ws["A2"].fill = HEADER_FILL

        ws["B1"].fill = HEADER_FILL
        ws["B2"].fill = HEADER_FILL

        ws["C1"].value = "Recherche de doublons de factures d'achats"
        ws["C1"].font = HEADER_FONT
        ws["C1"].fill = HEADER_FILL

        ws["C2"].font = HEADER_FONT
        ws["C2"].fill = HEADER_FILL

        ws["D1"].fill = HEADER_FILL
        ws["D2"].fill = HEADER_FILL

        ws["D1"].value = "JIC le"
        ws["D1"].font = HEADER_FONT
        ws["D1"].fill = HEADER_FILL

        ws["D2"].font = HEADER_FONT
        ws["D2"].fill = HEADER_FILL

    def create_objectifs(self, ws):
        ws["A5"].value = "Objectifs"
        ws["A5"].font = HEADER_FONT
        ws["A5"].fill = HEADER_FILL

        ws["A6"].value = "Lutte contre le détournement"

    def create_travaux_realise(self, ws):
        ws["A8"].value = "Travaux réalisés :"
        ws["A8"].font = HEADER_FONT
        ws["A8"].fill = HEADER_FILL

        ws["A9"].value = "A partir du FEC, filtre des comptes #6 et recherche de doublons via les critères suivants :"
        ws["A10"].value = "     - Comptenum"
        ws["A11"].value = "     - Ecrituredate"
        ws["A12"].value = "     - Pieceref"
        ws["A13"].value = "     - Ecriturelib"
        ws["A14"].value = "     - Solde"

        ws["A17"].value = "Racine"
        ws["A17"].font = HEADER_FONT
        ws["A17"].fill = HEADER_FILL

        ws["B17"].value = "6"
        ws["B17"].font = HEADER_FONT
        ws["B17"].fill = HEADER_FILL

    def clean_illegal_chars(self) -> pd.DataFrame:
        """Supprime les caractères de contrôle illégaux pour openpyxl dans les colonnes texte."""
        for col in self.df.select_dtypes(include='object').columns:
            self.df[col] = self.df[col].astype(str).apply(
                lambda x: ILLEGAL_CHARS_RE.sub('', x) if isinstance(x, str) else x
            )
        return self.df

    def add_racine_col(self):
        self.df.insert(5, "Racine", self.df["CompteNum"].astype(str).str[0])

    def add_solde_Col(self):
        debit = self.df["Debit"].astype(str).str.replace(',', '.', regex=False).astype(float)
        credit = self.df["Credit"].astype(str).str.replace(',', '.', regex=False).astype(float)
        self.df.insert(14, 'Solde', debit - credit)

    def search_all_six(self) -> pd.DataFrame:
        df_cop = self.df[self.df['Racine'] == '6'].copy()
        df_cop['EcritureDate'] = pd.to_datetime(df_cop['EcritureDate'], format='%Y%m%d').dt.strftime('%d/%m/%y')  # type: ignore[union-attr]

        grouped = (
            df_cop.groupby(['CompteNum', 'CompteLib', 'EcritureDate', 'EcritureLib', 'Solde'])
            .size()
            .reset_index(name='Nombre_Ecritures')  # type: ignore[call-overload]
        )
        result_df = grouped[grouped['Nombre_Ecritures'] >= 2].copy()

        if result_df.empty:
            print("Aucune écriture en double détectée")
            return pd.DataFrame(columns=pd.Index(['CompteNum', 'CompteLib', 'EcritureDate', 'Montant', 'Nombre_Ecritures']))

        result_df = result_df.sort_values(by=['CompteNum', 'EcritureDate']).reset_index(drop=True)  # type: ignore[call-overload]

        print(f"Nombre d'écritures en double détectées: {len(result_df)}")
        return result_df

    def create_df_tab(self, ws, df: pd.DataFrame):
        # Création du Header du tableau
        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=19, column=col)
            cell.value = header
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if header == 'Commentaire CAC':
                cell.font = HEADER_FONT_RED
            else:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT

        # Colonnes du df à insérer (dans l'ordre des HEADERS, sans "Commentaire CAC")
        df_columns = ['CompteNum', 'CompteLib', 'EcritureDate', 'EcritureLib', 'Solde', 'Nombre_Ecritures']

        for row_idx, (_, row) in enumerate(df[df_columns].iterrows(), start=20):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ligne de total
        total_row = 20 + len(df)

        for col in range(1, len(HEADERS)):
            cell = ws.cell(row=total_row, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if col == 1:
                cell.value = "Total"
            if col == 6:
                cell.value = df['Nombre_Ecritures'].sum()

    def run(self, output_path: str = 'Res.xlsx'):
        self.set_final_fec_date()
        self.open_fec()
        self.add_racine_col()
        self.add_solde_Col()
        print(self.df)

        occurence = self.search_all_six()
        print(occurence)

        self.clean_illegal_chars()

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            wb = writer.book
            wb.create_sheet(f"FEC {self.group_name}", 1)
            self.df.to_excel(writer, sheet_name=f"FEC {self.group_name}", index=False)
            ws = wb.create_sheet("Recherche Doublon", 0)
            self.create_entete(ws)
            self.create_objectifs(ws)
            self.create_travaux_realise(ws)
            self.create_df_tab(ws, occurence)